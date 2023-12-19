# coding=utf-8
import numpy as np
import pandas as pd

import os

from sklearn.utils import resample

from configuration_file import configuration_file
from openpyxl import Workbook
from rankboostYin import *
import xlrd

class Processing():
    def __init__(self):
        self.folder_name = configuration_file().datafolderPath

    def import_data(self):

        dataset = pd.core.frame.DataFrame()

        # In Mac the path use '/' to identify the secondary path
        folder_path = self.folder_name

        for root, dirs, files in os.walk(folder_path):

            for file in files:
                file_path = os.path.join(root, file)

                data1 = pd.read_csv(file_path)

                dataset = dataset.append(data1, ignore_index=True)

        return dataset

    def import_single_data(self):

        dataset = pd.core.frame.DataFrame()

        folder_path = self.folder_name

        for root, dirs, files in os.walk(folder_path):

            for file in files:
                file_path = os.path.join(root, file)

                dataset = pd.read_csv(file_path)
                yield dataset, file

    def separate_data(self, original_data):

        original_data = original_data.iloc[:, 3:]

        original_data = np.array(original_data)

        print("筛选掉0行之前的原始数据shape为:{0}".format(original_data.shape))
        filter_list = []
        for each in original_data:
            if int(each[10]) != 0:
                filter_list.append(each.tolist())
        original_data = np.array(filter_list)
        print("筛选掉0行之后的原始数据shape为:{0}".format(original_data.shape))


        training_data = resample(original_data)

        k = len(training_data[0])


        original_data = original_data.tolist()

        training_data = training_data.tolist()

        testing_data = []

        for i in original_data:
            if i not in training_data:
                testing_data.append(i)

        testing_data = np.array(testing_data)
        training_data = np.array(training_data)
        training_data_X = training_data[:, 0:k - 1]
        training_data_y = training_data[:, k - 1]
        testing_data_X = testing_data[:, 0:k - 1]
        testing_data_y = testing_data[:, k - 1]

        return training_data_X, training_data_y, testing_data_X, testing_data_y

    def split_train_test_csv(self, dataset, filename):
        count = 0
        for _ in range(100000000):
            try:
                print("{}数据集第{}次bootstrap".format(filename, _))
                training_data_X, training_data_y, testing_data_X, testing_data_y = Processing().separate_data(dataset)
                print(training_data_X.shape, training_data_y.shape, testing_data_X.shape, testing_data_y.shape)
                Cla_training_data_y = [1 if y > 0 else 0 for y in training_data_y]
                Cla_testing_data_y = [1 if y > 0 else 0 for y in testing_data_y]
                codeN = [i[10] for i in testing_data_X]


                if np.sum(Cla_training_data_y) == 0 or len(Cla_training_data_y) == np.sum(
                        Cla_training_data_y) or np.sum(Cla_testing_data_y) == 0 or \
                                len(Cla_testing_data_y) == np.sum(Cla_testing_data_y):
                    continue

                if len(np.shape(np.array(training_data_X))) != 2 or len(
                        np.shape(np.array(training_data_y))) != 1 or len(
                        np.shape(np.array(testing_data_X))) != 2 or len(np.shape(np.array(testing_data_y))) != 1:
                    continue

                count += 1


                wb = Workbook()
                ws = wb.active

                for i in range(len(training_data_X)):
                    l = []
                    l.extend(training_data_X[i])
                    l.append(training_data_y[i])
                    ws.append(l)
                dir_name = "".join(filename.split(".")[:-1])
                dir_name = os.path.join(configuration_file().bootstrap_dir, dir_name)
                train_csv_name = "".join(filename.split(".")[:-1]) + "_train_" + str(count) + ".xlsx"
                mkdir(dir_name)
                save_path = os.path.join(dir_name, train_csv_name)
                wb.save(save_path)


                wb_ = Workbook()
                ws_ = wb_.active
                for i in range(len(testing_data_X)):
                    l = []
                    l.extend(testing_data_X[i])
                    l.append(testing_data_y[i])
                    ws_.append(l)
                dir_name = "".join(filename.split(".")[:-1])
                dir_name = os.path.join(configuration_file().bootstrap_dir, dir_name)
                train_csv_name = "".join(filename.split(".")[:-1]) + "_test_" + str(count) + ".xlsx"
                mkdir(dir_name)
                save_path = os.path.join(dir_name, train_csv_name)
                wb_.save(save_path)

            except BaseException as BE:
                print("error type", BE)
                print(filename + "error")
            finally:
                if count == configuration_file().bootstrap_count:
                    print("bootstrap number is ：{}times".format(configuration_file().bootstrap_count))
                    break
        pass

    def read_bootstrap_csv(self):

        folder = configuration_file().bootstrap_dir
        datafolderPath = configuration_file().datafolderPath
        csv_list = os.listdir(datafolderPath)
        for csv_file in csv_list:
            name = "".join(csv_file.split(".")[:-1])
            train_data_x_list = []
            train_data_y_list = []
            test_data_x_list = []
            test_data_y_list = []
            for i in range(1, configuration_file().bootstrap_count + 1):
                tmp_train_x = []
                tmp_train_y = []
                tmp_test_x = []
                tmp_test_y = []

                bootstrap_train_name = name + "_train_" + str(i) + ".xlsx"
                bootstrap_test_name = name + "_test_" + str(i) + ".xlsx"
                bootstrap_train_ = os.path.join(folder, name)
                bootstrap_train_path = os.path.join(bootstrap_train_, bootstrap_train_name)
                bootstrap_test_path = os.path.join(bootstrap_train_, bootstrap_test_name)

                if not os.path.exists(bootstrap_train_path):
                    break
                if not os.path.exists(bootstrap_test_path):
                    break
                # 读train
                data = xlrd.open_workbook(bootstrap_train_path)

                sheeti = data.sheets()[0]
                nrows = sheeti.nrows
                ncols = sheeti.ncols

                for row in range(nrows):
                    row_list = []
                    tmp_train_y.append(sheeti.cell(row, ncols - 1).value)
                    for col in range(ncols - 1):
                        row_list.append(sheeti.cell(row, col).value)
                    tmp_train_x.append(row_list)
                train_data_x_list.append(tmp_train_x)
                train_data_y_list.append(tmp_train_y)

                # 读test
                data = xlrd.open_workbook(bootstrap_test_path)

                sheeti = data.sheets()[0]
                nrows = sheeti.nrows
                ncols = sheeti.ncols

                for row in range(nrows):
                    row_list = []
                    tmp_test_y.append(sheeti.cell(row, ncols - 1).value)
                    for col in range(ncols - 1):
                        row_list.append(sheeti.cell(row, col).value)
                    tmp_test_x.append(row_list)
                test_data_x_list.append(tmp_test_x)
                test_data_y_list.append(tmp_test_y)
            yield train_data_x_list, train_data_y_list, test_data_x_list, test_data_y_list, csv_file

    def import_crossversion_data(self):

        dataset_train = pd.core.frame.DataFrame()
        dataset_test = pd.core.frame.DataFrame()

        folder_path = configuration_file().crossversiondatafolderPath + '/'

        def transform_data(original_data):
            original_data = original_data.iloc[:, :]

            original_data = np.array(original_data)

            k = len(original_data[0])


            original_data = sorted(
                original_data, key=lambda x: x[-1], reverse=True)

            original_data = np.array(original_data)
            original_data_X = original_data[:, 0:k - 1]

            original_data_y = original_data[:, k - 1]

            return original_data_X, original_data_y

        for root, dirs, files, in os.walk(folder_path):
            if root == folder_path:

                thisroot = root
                for dir in dirs:
                    dir_path = os.path.join(thisroot, dir)

                    for root, dirs, files, in os.walk(dir_path):
                        if(files[0][-7:-4]<files[1][-7:-4]):
                            file_path_train = os.path.join(dir_path, files[0])
                            file_path_test = os.path.join(dir_path, files[1])
                            trainingfile=files[0]
                            testingfile=files[1]
                        else:
                            file_path_train = os.path.join(dir_path, files[1])
                            file_path_test = os.path.join(dir_path, files[0])
                            trainingfile = files[1]
                            testingfile = files[0]

                        print('files[0][-7:-4]', files[0][-7:-4])
                        print('files[1][-7:-4]', files[1][-7:-4])
                        print(files[0][-7:-4] > files[1][-7:-4])
                        print('train', file_path_train)
                        print('test', file_path_test)

                        dataset_train = pd.read_csv(file_path_train)
                        dataset_test = pd.read_csv(file_path_test)
                        training_data_x, training_data_y = transform_data(
                            dataset_train)
                        testing_data_x, testing_data_y = transform_data(
                            dataset_test)
                        yield training_data_x, training_data_y, testing_data_x, testing_data_y, dir, trainingfile, testingfile


    def write_excel(self, excel_path, data):

        dir_name = str(os.path.split(excel_path)[0])
        print(dir_name)
        mkdir(dir_name)
        wb = Workbook()
        ws = wb.active
        for _ in data:
            ws.append(_)
        wb.save(excel_path)


    def change_to_newdata(self, training_data_X, training_data_y, testing_data_X, testing_data_y):

        want_row = [i for i in range(len(training_data_X))]
        new_train_data_x = training_data_X[want_row]
        want_row = [i for i in range(len(testing_data_X))]
        new_test_data_x = testing_data_X[want_row]

        want_col = [j for j in range(0, 10)] + [j for j in range(11, len(training_data_X[0]))]
        new_train_data_x = new_train_data_x[:, want_col]
        new_test_data_x = new_test_data_x[:, want_col]

        loc_train = training_data_X[:, [10]].squeeze()
        loc_test = testing_data_X[:, [10]].squeeze()

        new_train_data_y = training_data_y / loc_train
        new_test_data_y = testing_data_y / loc_test
        return new_train_data_x, new_train_data_y, new_test_data_x, new_test_data_y

    def newdata(self, training_data_X, training_data_y, testing_data_X, testing_data_y, index):

        want_row = [i for i in range(len(training_data_X))]
        new_train_data_x = training_data_X[want_row]
        want_row = [i for i in range(len(testing_data_X))]
        new_test_data_x = testing_data_X[want_row]

        want_col = [j for j in range(0, index)] + [j for j in range(index + 1, len(training_data_X[0]))]
        print('wantcol',want_col)
        new_train_data_x = new_train_data_x[:, want_col]
        new_test_data_x = new_test_data_x[:, want_col]


        loc_train = training_data_X[:, [index]].squeeze()
        loc_test = testing_data_X[:, [index]].squeeze()
        new_train_data_y = [i for i in training_data_y]
        new_test_data_y = [i for i in testing_data_y]
        for i in range(len(training_data_y)):
            if (loc_train[i] > 0.0):
                new_train_data_y[i] = training_data_y[i] / loc_train[i]
            else:
                new_train_data_y[i] = 0

        for i in range(len(testing_data_y)):
            if (loc_test[i] > 0.0):
                new_test_data_y[i] = testing_data_y[i] / loc_test[i]
            else:
                new_test_data_y[i] = 0

        return new_train_data_x, new_train_data_y, new_test_data_x, new_test_data_y


    def newjitdata(self, training_data_X, training_data_y, testing_data_X, testing_data_y, laindex, ldindex):

        want_row = [i for i in range(len(training_data_X))]
        new_train_data_x = training_data_X[want_row]
        want_row = [i for i in range(len(testing_data_X))]
        new_test_data_x = testing_data_X[want_row]

        want_col = [j for j in range(0, laindex)] + [j for j in range(ldindex + 1, len(training_data_X[0]))]
        print('wantcol',want_col)
        new_train_data_x = new_train_data_x[:, want_col]
        new_test_data_x = new_test_data_x[:, want_col]


        loc_train = [i[laindex] + i[ldindex] for i in training_data_X]
        loc_test = [i[laindex] + i[ldindex] for i in testing_data_X]

        new_train_data_y = [i for i in training_data_y]
        new_test_data_y = [i for i in testing_data_y]


        for i in range(len(training_data_y)):
            if (loc_train[i] > 0.0):
                new_train_data_y[i] = training_data_y[i] / loc_train[i]
            else:
                new_train_data_y[i] = 0

        for i in range(len(testing_data_y)):
            if (loc_test[i] > 0.0):
                new_test_data_y[i] = testing_data_y[i] / loc_test[i]
            else:
                new_test_data_y[i] = 0

        return new_train_data_x, new_train_data_y, new_test_data_x, new_test_data_y